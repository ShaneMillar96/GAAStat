#!/usr/bin/env python3
"""
Detailed analysis of Excel file to understand the structure and find percentage issues
"""
import openpyxl
from openpyxl import load_workbook

def detailed_excel_analysis(file_path):
    """Perform detailed analysis of Excel structure and data"""
    try:
        workbook = load_workbook(file_path)
        print(f"📊 Detailed Excel Analysis: {file_path}")
        
        # Find player stats sheets
        player_stats_sheets = [name for name in workbook.sheetnames 
                              if "player" in name.lower() and "stats" in name.lower()]
        
        print(f"🏈 Analyzing {len(player_stats_sheets)} player stats sheets")
        
        # Take first available player stats sheet for detailed analysis
        if not player_stats_sheets:
            print("❌ No player stats sheets found")
            return
        
        sheet_name = player_stats_sheets[0]  # Use first sheet
        worksheet = workbook[sheet_name]
        
        print(f"\n📄 Detailed analysis of: {sheet_name}")
        print(f"   Dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns")
        
        # Show first few rows and columns to understand structure
        print(f"\n📋 Headers (Row 1):")
        headers = []
        for col in range(1, min(worksheet.max_column + 1, 21)):  # First 20 columns
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append((col, cell_value))
            print(f"   Column {col}: {cell_value}")
        
        print(f"\n📊 Sample Data (Rows 2-4):")
        for row in range(2, min(5, worksheet.max_row + 1)):
            print(f"   Row {row}:")
            for col in range(1, min(worksheet.max_column + 1, 11)):  # First 10 columns
                cell_value = worksheet.cell(row=row, column=col).value
                print(f"     Col {col}: {cell_value}")
        
        # Look for all numeric columns that might contain percentages
        print(f"\n🔍 Analyzing all numeric columns for potential percentage violations:")
        
        violation_found = False
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if not header:
                continue
                
            # Check values in this column
            values = []
            violations = []
            
            for row in range(2, min(worksheet.max_row + 1, 21)):  # Check first 20 rows
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    try:
                        value = float(cell.value)
                        values.append(value)
                        
                        # Check if this could be a percentage field that violates 0-1 range
                        if value < 0 or value > 1:
                            violations.append((row, value))
                    except (ValueError, TypeError):
                        pass
            
            # Only report columns with violations
            if violations:
                violation_found = True
                print(f"\n   ❌ Column {col} ({header}):")
                print(f"      Values range: {min(values):.4f} to {max(values):.4f}")
                print(f"      {len(violations)} violations of 0-1 range:")
                for row, val in violations[:5]:  # Show first 5
                    print(f"        Row {row}: {val}")
                if len(violations) > 5:
                    print(f"        ... and {len(violations) - 5} more")
        
        if not violation_found:
            print("   ✅ No obvious percentage violations found in numeric columns")
            
            # Let's check if there are any calculated percentage fields that might be the issue
            print(f"\n🧮 Checking for calculated percentage fields...")
            
            # Look for specific column patterns that might contain percentages
            percentage_keywords = [
                'efficiency', 'success', 'conversion', 'rate', 'percent', 
                'pct', 'accuracy', 'completion', 'ratio'
            ]
            
            potential_percentage_columns = []
            for col, header in headers:
                if header and any(keyword in str(header).lower() for keyword in percentage_keywords):
                    potential_percentage_columns.append((col, header))
            
            if potential_percentage_columns:
                print(f"   Found {len(potential_percentage_columns)} potential percentage columns:")
                for col, header in potential_percentage_columns:
                    print(f"     Column {col}: {header}")
                    
                    # Check values
                    for row in range(2, min(worksheet.max_row + 1, 6)):
                        cell_value = worksheet.cell(row=row, column=col).value
                        if cell_value is not None:
                            print(f"       Row {row}: {cell_value} (type: {type(cell_value).__name__})")
            else:
                print("   No obvious percentage column names found")
        
        return violation_found
        
    except Exception as e:
        print(f"❌ Error in detailed analysis: {e}")
        return False

if __name__ == "__main__":
    file_path = "/Users/shane.millar/Downloads/Drum Analysis 2025.xlsx"
    detailed_excel_analysis(file_path)