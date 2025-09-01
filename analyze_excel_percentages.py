#!/usr/bin/env python3
"""
Analyze Excel file to identify percentage values and potential constraint violations
"""
import sys
import openpyxl
from openpyxl import load_workbook
import re

def analyze_excel_percentages(file_path):
    """Analyze Excel file for percentage values that might violate database constraints"""
    try:
        workbook = load_workbook(file_path)
        print(f"📊 Analyzing Excel file: {file_path}")
        print(f"📋 Found {len(workbook.sheetnames)} sheets")
        
        # Focus on player stats sheets
        player_stats_sheets = []
        for sheet_name in workbook.sheetnames:
            if "player" in sheet_name.lower() and "stats" in sheet_name.lower():
                player_stats_sheets.append(sheet_name)
        
        print(f"🏈 Found {len(player_stats_sheets)} player stats sheets:")
        for sheet in player_stats_sheets:
            print(f"  - {sheet}")
        
        # Analyze each player stats sheet
        percentage_issues = []
        
        for sheet_name in player_stats_sheets:
            worksheet = workbook[sheet_name]
            print(f"\n📈 Analyzing sheet: {sheet_name}")
            
            # Look for percentage columns (commonly named columns)
            percentage_columns = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    cell_str = str(cell_value).lower()
                    if any(keyword in cell_str for keyword in ['efficiency', 'success', 'conversion', 'rate', '%']):
                        percentage_columns.append((col, cell_value))
            
            print(f"  Found {len(percentage_columns)} potential percentage columns:")
            for col, header in percentage_columns:
                print(f"    Column {col}: {header}")
            
            # Check values in percentage columns
            for col, header in percentage_columns:
                print(f"\n  🔍 Analyzing column {col} ({header}):")
                values = []
                invalid_values = []
                
                for row in range(2, min(worksheet.max_row + 1, 52)):  # Check first 50 data rows
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        try:
                            value = float(cell.value)
                            values.append(value)
                            
                            # Check if value violates constraint (should be 0-1)
                            if value < 0 or value > 1:
                                invalid_values.append((row, value))
                        except (ValueError, TypeError):
                            # Non-numeric values
                            print(f"    Row {row}: Non-numeric value '{cell.value}'")
                
                if values:
                    print(f"    📊 Found {len(values)} numeric values")
                    print(f"    📏 Range: {min(values):.4f} to {max(values):.4f}")
                    print(f"    📈 Average: {sum(values)/len(values):.4f}")
                    
                    if invalid_values:
                        print(f"    ❌ {len(invalid_values)} values violate constraint (not 0-1):")
                        for row, val in invalid_values[:10]:  # Show first 10 violations
                            print(f"      Row {row}: {val}")
                        if len(invalid_values) > 10:
                            print(f"      ... and {len(invalid_values) - 10} more")
                        
                        percentage_issues.extend([(sheet_name, header, row, val) for row, val in invalid_values])
                    else:
                        print(f"    ✅ All values are within valid range (0-1)")
        
        # Summary
        print(f"\n📋 SUMMARY:")
        print(f"Total percentage constraint violations: {len(percentage_issues)}")
        
        if percentage_issues:
            print(f"\n❌ CONSTRAINT VIOLATIONS FOUND:")
            sheets_with_issues = {}
            for sheet, header, row, val in percentage_issues:
                if sheet not in sheets_with_issues:
                    sheets_with_issues[sheet] = {}
                if header not in sheets_with_issues[sheet]:
                    sheets_with_issues[sheet][header] = []
                sheets_with_issues[sheet][header].append((row, val))
            
            for sheet, columns in sheets_with_issues.items():
                print(f"\n  📄 {sheet}:")
                for header, violations in columns.items():
                    print(f"    🔸 {header}: {len(violations)} violations")
                    print(f"      Range: {min(v[1] for v in violations):.4f} to {max(v[1] for v in violations):.4f}")
            
            print(f"\n💡 RECOMMENDATIONS:")
            print(f"1. Values > 1 appear to be percentages (0-100) that need division by 100")
            print(f"2. Check Excel parsing logic in CreatePlayerStatFromData method")
            print(f"3. Add percentage normalization for columns: {set(h for _, h, _, _ in percentage_issues)}")
        else:
            print(f"✅ No percentage constraint violations found")
        
        return percentage_issues
        
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        return []

if __name__ == "__main__":
    file_path = "/Users/shane.millar/Downloads/Drum Analysis 2025.xlsx"
    issues = analyze_excel_percentages(file_path)
    
    if issues:
        sys.exit(1)  # Exit with error if issues found
    else:
        sys.exit(0)  # Exit success